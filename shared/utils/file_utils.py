"""
File Utilities

Provides Excel I/O operations for data persistence.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import logging

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

logger = logging.getLogger(__name__)


def read_excel(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Read Excel file and return list of dictionaries.
    
    First row is treated as header with column names.
    
    Args:
        path: Path to Excel file
        sheet_name: Optional sheet name (defaults to active sheet)
        
    Returns:
        List of dictionaries, one per row
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ImportError: If openpyxl is not installed
        
    Example:
        >>> records = read_excel("data/file_1_2.xlsx")
        >>> print(records[0]['email_id'])
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required for Excel operations")
    
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    
    wb = load_workbook(filename=path, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    
    # First row is header
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    
    records = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue  # Skip empty rows
        record = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers)}
        records.append(record)
    
    wb.close()
    logger.debug(f"Read {len(records)} records from {path}")
    return records


def write_excel(
    path: str, 
    records: List[Dict[str, Any]], 
    columns: Optional[List[str]] = None,
    sheet_name: str = "Sheet1"
) -> None:
    """
    Write list of dictionaries to Excel file.
    
    Creates parent directories if they don't exist.
    
    Args:
        path: Path to output Excel file
        records: List of dictionaries to write
        columns: Optional list of columns to include (and their order)
        sheet_name: Name for the worksheet
        
    Raises:
        ImportError: If openpyxl is not installed
        
    Example:
        >>> records = [{"email_id": "abc", "grade": 85}]
        >>> write_excel("output/grades.xlsx", records)
    """
    if Workbook is None:
        raise ImportError("openpyxl is required for Excel operations")
    
    if not records:
        logger.warning(f"No records to write to {path}")
        return
    
    # Ensure directory exists
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Determine columns
    if columns is None:
        # Use keys from first record
        columns = list(records[0].keys())
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = record.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(path)
    logger.info(f"Wrote {len(records)} records to {path}")
