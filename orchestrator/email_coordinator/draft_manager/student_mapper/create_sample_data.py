#!/usr/bin/env python3
"""
Script to create sample students_mapping.xlsx file
"""
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "Students"

# Add headers
ws['A1'] = 'email_address'
ws['B1'] = 'name'

# Add sample student data
students = [
    ('student1@example.com', 'Alex Johnson'),
    ('student2@example.com', 'Maria Garcia'),
    ('student3@example.com', 'James Smith'),
    ('student4@example.com', 'Sarah Williams'),
    ('student5@example.com', 'Michael Brown'),
    ('test@university.edu', 'Test Student'),
    ('john.doe@college.edu', 'John Doe'),
    ('jane.smith@university.edu', 'Jane Smith'),
]

for idx, (email, name) in enumerate(students, start=2):
    ws[f'A{idx}'] = email
    ws[f'B{idx}'] = name

# Save the file
wb.save('./data/students_mapping.xlsx')
print("Sample data file created successfully at ./data/students_mapping.xlsx")
