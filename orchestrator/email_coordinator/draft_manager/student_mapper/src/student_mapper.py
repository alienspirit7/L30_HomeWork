"""
Student Mapper Service
Maps email addresses to student names using an Excel lookup table.
"""
import logging
import os
from pathlib import Path
from typing import Dict, Optional
from openpyxl import load_workbook
import yaml


class StudentMapper:
    """
    Service to map email addresses to student names.

    Features:
    - Load student mapping from Excel file
    - In-memory caching for performance
    - Case-insensitive email matching
    - Fallback name for unknown students
    """

    def __init__(self, config_path: str = "./config.yaml"):
        """
        Initialize the Student Mapper service.

        Args:
            config_path: Path to the configuration file
        """
        self.config = self._load_config(config_path)
        self.logger = self._setup_logging()
        self.mapping_cache: Dict[str, str] = {}
        self._load_mapping()

    def _load_config(self, config_path: str) -> dict:
        """Load configuration from YAML file."""
        with open(config_path, 'r') as f:
            return yaml.safe_load(f)

    def _setup_logging(self) -> logging.Logger:
        """Setup logging configuration."""
        logger = logging.getLogger(self.config['service']['name'])
        logger.setLevel(self.config['logging']['level'])

        # Create logs directory if it doesn't exist
        log_file = self.config['logging']['file']
        log_dir = os.path.dirname(log_file)
        if log_dir:
            os.makedirs(log_dir, exist_ok=True)

        # File handler
        fh = logging.FileHandler(log_file)
        fh.setLevel(self.config['logging']['level'])

        # Console handler
        ch = logging.StreamHandler()
        ch.setLevel(self.config['logging']['level'])

        # Formatter
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)

        logger.addHandler(fh)
        logger.addHandler(ch)

        return logger

    def _load_mapping(self) -> None:
        """
        Load student email-to-name mapping from Excel file.

        Raises:
            FileNotFoundError: If mapping file doesn't exist
        """
        mapping_file = self.config['data']['mapping_file']

        if not os.path.exists(mapping_file):
            raise FileNotFoundError(f"Mapping file not found: {mapping_file}")

        self.logger.info(f"Loading student mapping from {mapping_file}")

        try:
            wb = load_workbook(mapping_file, read_only=True)
            ws = wb.active

            email_col = self.config['columns']['email_column']
            name_col = self.config['columns']['name_column']

            # Find column indices
            headers = [cell.value for cell in ws[1]]
            email_idx = headers.index(email_col) + 1
            name_idx = headers.index(name_col) + 1

            # Load data (skip header row)
            for row in ws.iter_rows(min_row=2, values_only=True):
                email = row[email_idx - 1]
                name = row[name_idx - 1]

                if email and name:
                    # Store with lowercase email for case-insensitive lookup
                    self.mapping_cache[email.lower().strip()] = name.strip()

            wb.close()

            self.logger.info(
                f"Loaded {len(self.mapping_cache)} student mappings"
            )

        except Exception as e:
            self.logger.error(f"Error loading mapping file: {e}")
            raise

    def map_email_to_name(self, email_address: str) -> Dict[str, any]:
        """
        Map an email address to a student name.

        Args:
            email_address: The email address to look up

        Returns:
            dict: {
                "name": str | None,
                "found": bool
            }
        """
        if not email_address:
            self.logger.warning("Empty email address provided")
            return {
                "name": self.config['defaults']['fallback_name'],
                "found": False
            }

        # Case-insensitive lookup
        email_lower = email_address.lower().strip()

        if email_lower in self.mapping_cache:
            name = self.mapping_cache[email_lower]
            self.logger.debug(f"Found mapping: {email_address} -> {name}")
            return {
                "name": name,
                "found": True
            }
        else:
            fallback = self.config['defaults']['fallback_name']
            self.logger.debug(
                f"No mapping found for {email_address}, using fallback: {fallback}"
            )
            return {
                "name": fallback,
                "found": False
            }

    def reload_mapping(self) -> None:
        """Reload the mapping from the Excel file."""
        self.logger.info("Reloading student mapping")
        self.mapping_cache.clear()
        self._load_mapping()

    def get_stats(self) -> Dict[str, any]:
        """
        Get statistics about the loaded mappings.

        Returns:
            dict: Statistics including total mappings count
        """
        return {
            "total_mappings": len(self.mapping_cache),
            "service": self.config['service']['name'],
            "version": self.config['service']['version']
        }


# Main interface function for parent service to call
def lookup_student(email_address: str, config_path: str = "./config.yaml") -> Dict[str, any]:
    """
    Standalone function to lookup a student name by email.

    Args:
        email_address: Email to look up
        config_path: Path to config file

    Returns:
        dict: {"name": str | None, "found": bool}
    """
    mapper = StudentMapper(config_path)
    return mapper.map_email_to_name(email_address)
