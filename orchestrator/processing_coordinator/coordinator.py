"""
Processing Coordinator
Level 1 - Domain Coordinator

Coordinates repository processing operations:
- Grading: Cloning repos and analyzing code (via grade_manager)
- Feedback: Generating AI feedback (via feedback_manager)
"""

import sys
import os
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
import yaml

# Add parent path for imports
sys.path.insert(0, str(Path(__file__).parent))

# Import child managers
try:
    from grade_manager.service import GradeManagerService
except ImportError:
    GradeManagerService = None

try:
    from feedback_manager.service import FeedbackManager
except ImportError:
    FeedbackManager = None

logger = logging.getLogger(__name__)


class ProcessingCoordinator:
    """
    Processing Coordinator - Level 1 Domain Coordinator
    
    Coordinates processing operations by delegating to child managers:
    - Grade Manager: Clones repos and calculates grades
    - Feedback Manager: Generates AI-powered feedback
    
    Actions:
    - grade: Clone and grade repositories
    - feedback: Generate AI feedback for grades
    - full_pipeline: Complete processing workflow
    """
    
    def __init__(self, config_path: str = "config.yaml"):
        """
        Initialize the Processing Coordinator.
        
        Args:
            config_path: Path to configuration file
        """
        self.config_path = config_path
        self.config = self._load_config(config_path)
        self._setup_logging()
        self._initialize_child_managers()
        
        self.logger.info("Processing Coordinator initialized")
    
    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load configuration from YAML file."""
        try:
            with open(config_path, 'r') as f:
                return yaml.safe_load(f) or {}
        except FileNotFoundError:
            logger.warning(f"Config file not found: {config_path}")
            return self._default_config()
    
    def _default_config(self) -> Dict[str, Any]:
        """Return default configuration."""
        return {
            'coordinator': {
                'name': 'processing_coordinator',
                'version': '1.0.0'
            },
            'children': {
                'grade_manager': './grade_manager',
                'feedback_manager': './feedback_manager'
            },
            'logging': {
                'level': 'INFO',
                'file': './logs/processing_coordinator.log'
            }
        }
    
    def _setup_logging(self) -> None:
        """Configure logging."""
        log_config = self.config.get('logging', {})
        log_level = getattr(logging, log_config.get('level', 'INFO').upper())
        log_file = log_config.get('file', './logs/processing_coordinator.log')
        
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level)
        
        if not self.logger.handlers:
            fh = logging.FileHandler(log_file)
            fh.setLevel(log_level)
            fh.setFormatter(logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            ))
            self.logger.addHandler(fh)
            
            ch = logging.StreamHandler()
            ch.setLevel(log_level)
            ch.setFormatter(logging.Formatter('%(levelname)s - %(message)s'))
            self.logger.addHandler(ch)
    
    def _initialize_child_managers(self) -> None:
        """Initialize child managers."""
        children_config = self.config.get('children', {})
        base_path = Path(self.config_path).parent
        
        # Initialize Grade Manager
        if GradeManagerService is not None:
            grade_path = base_path / children_config.get('grade_manager', './grade_manager')
            grade_config = grade_path / 'config.yaml'
            try:
                self.grade_manager = GradeManagerService(
                    config_path=str(grade_config) if grade_config.exists() else "config.yaml"
                )
                self.logger.info("Grade Manager initialized")
            except Exception as e:
                self.logger.warning(f"Failed to initialize Grade Manager: {e}")
                self.grade_manager = None
        else:
            self.grade_manager = None
        
        # Initialize Feedback Manager
        if FeedbackManager is not None:
            feedback_path = base_path / children_config.get('feedback_manager', './feedback_manager')
            feedback_config = feedback_path / 'config.yaml'
            try:
                self.feedback_manager = FeedbackManager(
                    config_path=str(feedback_config) if feedback_config.exists() else "config.yaml"
                )
                self.logger.info("Feedback Manager initialized")
            except Exception as e:
                self.logger.warning(f"Failed to initialize Feedback Manager: {e}")
                self.feedback_manager = None
        else:
            self.feedback_manager = None
    
    def grade(self, email_records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Grade repositories from email records.
        
        Args:
            email_records: List of email records with repo_url
            
        Returns:
            Dictionary with grades, output_file, counts
        """
        self.logger.info(f"Grading {len(email_records)} repositories")
        
        if self.grade_manager is None:
            return {
                'grades': [],
                'output_file': None,
                'graded_count': 0,
                'failed_count': len(email_records),
                'status': 'failed',
                'error': 'Grade Manager not available'
            }
        
        try:
            result = self.grade_manager.process({
                'email_records': email_records
            })
            
            result['status'] = 'success'
            return result
            
        except Exception as e:
            self.logger.error(f"Error grading repositories: {e}")
            return {
                'grades': [],
                'output_file': None,
                'graded_count': 0,
                'failed_count': len(email_records),
                'status': 'failed',
                'error': str(e)
            }
    
    def generate_feedback(self, grade_records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate AI feedback for grade records.
        
        Args:
            grade_records: List of grade records
            
        Returns:
            Dictionary with feedback list and counts
        """
        self.logger.info(f"Generating feedback for {len(grade_records)} grades")
        
        if self.feedback_manager is None:
            return {
                'feedback': [],
                'output_file': None,
                'generated_count': 0,
                'failed_count': len(grade_records),
                'status': 'failed',
                'error': 'Feedback Manager not available'
            }
        
        try:
            result = self.feedback_manager.process(grade_records)
            
            result['status'] = 'success'
            return result
            
        except Exception as e:
            self.logger.error(f"Error generating feedback: {e}")
            return {
                'feedback': [],
                'output_file': None,
                'generated_count': 0,
                'failed_count': len(grade_records),
                'status': 'failed',
                'error': str(e)
            }
    
    def process(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process input based on action parameter.
        
        Args:
            input_data: Dictionary with 'action' and action-specific params
            
        Returns:
            Action-specific result dictionary
        """
        action = input_data.get('action', 'grade')
        
        if action == 'grade':
            return self.grade(
                email_records=input_data.get('email_records', [])
            )
        
        elif action == 'feedback':
            return self.generate_feedback(
                grade_records=input_data.get('grade_records', [])
            )
        
        elif action == 'full_pipeline':
            # Step 1: Grade repositories
            email_records = input_data.get('email_records', [])
            grade_result = self.grade(email_records)
            
            if grade_result.get('status') == 'failed':
                return {
                    'action': 'full_pipeline',
                    'grade_result': grade_result,
                    'feedback_result': None,
                    'status': 'failed',
                    'error': grade_result.get('error')
                }
            
            # Step 2: Generate feedback
            grade_records = grade_result.get('grades', [])
            feedback_result = self.generate_feedback(grade_records)
            
            return {
                'action': 'full_pipeline',
                'grade_result': grade_result,
                'feedback_result': feedback_result,
                'status': 'success' if feedback_result.get('status') == 'success' else 'partial'
            }
        
        else:
            return {
                'status': 'failed',
                'error': f'Unknown action: {action}'
            }
    
    def health_check(self) -> Dict[str, Any]:
        """Check coordinator and child manager health."""
        return {
            'coordinator': 'processing_coordinator',
            'status': 'healthy',
            'children': {
                'grade_manager': 'healthy' if self.grade_manager else 'unavailable',
                'feedback_manager': 'healthy' if self.feedback_manager else 'unavailable'
            }
        }


def main():
    """Main entry point for standalone execution."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Processing Coordinator')
    parser.add_argument('--action', choices=['grade', 'feedback', 'health'],
                       default='health', help='Action to perform')
    parser.add_argument('--input', help='Path to input Excel file')
    parser.add_argument('--config', default='config.yaml',
                       help='Path to config file')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Processing Coordinator v1.0.0")
    print("=" * 60)
    
    coordinator = ProcessingCoordinator(config_path=args.config)
    
    if args.action == 'health':
        health = coordinator.health_check()
        print(f"Status: {health['status']}")
        for child, status in health.get('children', {}).items():
            icon = "✓" if status == "healthy" else "✗"
            print(f"  {icon} {child}: {status}")
    
    elif args.action == 'grade' and args.input:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.input)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = rows[0]
            records = [dict(zip(headers, row)) for row in rows[1:] if any(row)]
            
            result = coordinator.grade(records)
            print(f"Status: {result.get('status')}")
            print(f"Graded: {result.get('graded_count', 0)}")
            print(f"Failed: {result.get('failed_count', 0)}")
        except Exception as e:
            print(f"Error: {e}")
    
    print("=" * 60)
    return 0


if __name__ == '__main__':
    sys.exit(main())
