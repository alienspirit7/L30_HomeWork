"""
Feedback Manager Service
Level 2 - Task Manager

Manages AI feedback generation: selecting feedback styles based on grades
and generating personalized feedback using Gemini API.
"""

import argparse
import logging
import os
import sys
import time
from pathlib import Path
from typing import Dict, Any, List, Optional

import yaml
from openpyxl import load_workbook, Workbook


class FeedbackManager:
    """
    Task Manager service that orchestrates feedback generation.
    Coordinates style_selector and gemini_generator services.
    """

    def __init__(self, config_path: str = "config.yaml"):
        """
        Initialize the Feedback Manager service.

        Args:
            config_path: Path to configuration file
        """
        self.config = self._load_config(config_path)
        self._setup_logging()
        self._initialize_child_services()

    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load configuration from YAML file."""
        try:
            config_file = Path(config_path)
            if not config_file.exists():
                raise FileNotFoundError(f"Config file not found: {config_path}")

            with open(config_file, 'r') as f:
                return yaml.safe_load(f)
        except yaml.YAMLError as e:
            raise ValueError(f"Invalid YAML in config file: {e}")

    def _setup_logging(self):
        """Configure logging based on config settings."""
        log_config = self.config.get('logging', {})
        log_level = getattr(logging, log_config.get('level', 'INFO'))
        log_file = log_config.get('file', './logs/feedback_manager.log')
        log_format = log_config.get('format',
            '%(asctime)s | %(levelname)s | %(name)s | %(message)s')

        # Create logs directory if it doesn't exist
        log_dir = Path(log_file).parent
        log_dir.mkdir(parents=True, exist_ok=True)

        # Setup handlers
        handlers = [logging.FileHandler(log_file)]
        if log_config.get('console', True):
            handlers.append(logging.StreamHandler(sys.stdout))

        logging.basicConfig(
            level=log_level,
            format=log_format,
            handlers=handlers
        )

        self.logger = logging.getLogger(self.config['manager']['name'])
        self.logger.info(
            f"Feedback Manager v{self.config['manager']['version']} initialized"
        )

    def _initialize_child_services(self):
        """Initialize child services (style_selector and gemini_generator)."""
        try:
            # Import child services
            sys.path.insert(0, str(Path(__file__).parent))

            from style_selector.service import StyleSelector
            from gemini_generator.service import GeminiGeneratorService

            # Initialize services
            style_selector_path = Path(self.config['children']['style_selector'])
            gemini_generator_path = Path(self.config['children']['gemini_generator'])

            style_config = style_selector_path / "config.yaml"
            gemini_config = gemini_generator_path / "config.yaml"

            self.style_selector = StyleSelector(config_path=str(style_config))
            self.gemini_generator = GeminiGeneratorService(config_path=str(gemini_config))

            self.logger.info("Child services initialized successfully")
        except Exception as e:
            self.logger.error(f"Failed to initialize child services: {e}")
            raise

    def _load_grade_records(self, input_file: str) -> List[Dict[str, Any]]:
        """
        Load grade records from Excel file.

        Args:
            input_file: Path to input Excel file

        Returns:
            List of grade records with status="Ready"
        """
        self.logger.info(f"Loading grade records from: {input_file}")

        try:
            wb = load_workbook(input_file)
            ws = wb.active

            # Get column mappings
            col_config = self.config['input']['columns']
            email_col = col_config['email_id']
            grade_col = col_config['grade']
            status_col = col_config['status']

            # Read header row to find column indices
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

            # Validate required columns exist
            required_cols = [email_col, grade_col, status_col]
            for col in required_cols:
                if col not in headers:
                    raise ValueError(f"Required column '{col}' not found in input file")

            # Extract records with status="Ready"
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                email_id = row[headers[email_col] - 1]
                grade = row[headers[grade_col] - 1]
                status = row[headers[status_col] - 1]

                if status == "Ready":
                    records.append({
                        'email_id': email_id,
                        'grade': float(grade) if grade is not None else None,
                        'status': status
                    })

            self.logger.info(f"Loaded {len(records)} records with status='Ready'")
            return records

        except FileNotFoundError:
            self.logger.error(f"Input file not found: {input_file}")
            raise
        except Exception as e:
            self.logger.error(f"Error loading grade records: {e}")
            raise

    def _generate_feedback_for_record(self, record: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generate feedback for a single grade record.

        Args:
            record: Grade record with email_id and grade

        Returns:
            Feedback record with email_id, reply, and status
        """
        email_id = record['email_id']
        grade = record['grade']

        try:
            # Step 1: Select style based on grade
            style_result = self.style_selector.process({'grade': grade})
            style_name = style_result['style_name']
            prompt = style_result['prompt_template']

            self.logger.debug(f"Selected style '{style_name}' for grade {grade}")

            # Step 2: Generate feedback using Gemini
            gemini_input = {
                'prompt': prompt,
                'style': style_name,
                'context': {
                    'grade': grade,
                    'email_id': email_id
                }
            }

            feedback_result = self.gemini_generator.process(gemini_input)

            if feedback_result['status'] == 'Success' and feedback_result['feedback']:
                return {
                    'email_id': email_id,
                    'reply': feedback_result['feedback'],
                    'status': 'Ready'
                }
            else:
                # API failed, mark as missing reply
                error_msg = feedback_result.get('error', 'Unknown error')
                self.logger.warning(
                    f"Failed to generate feedback for {email_id}: {error_msg}"
                )
                return {
                    'email_id': email_id,
                    'reply': None,
                    'status': 'Missing: reply'
                }

        except Exception as e:
            self.logger.error(f"Error processing record {email_id}: {e}")
            return {
                'email_id': email_id,
                'reply': None,
                'status': 'Missing: reply'
            }

    def _write_feedback_records(self, feedback_records: List[Dict[str, Any]],
                                output_file: str):
        """
        Write feedback records to Excel file.

        Args:
            feedback_records: List of feedback records
            output_file: Path to output Excel file
        """
        self.logger.info(f"Writing feedback records to: {output_file}")

        try:
            # Create output directory if it doesn't exist
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Create workbook
            wb = Workbook()
            ws = wb.active

            # Get column mappings
            col_config = self.config['output']['columns']

            # Write header
            headers = [col_config['email_id'], col_config['reply'], col_config['status']]
            ws.append(headers)

            # Write data
            for record in feedback_records:
                ws.append([
                    record['email_id'],
                    record['reply'],
                    record['status']
                ])

            # Save workbook
            wb.save(output_file)
            self.logger.info(f"Successfully wrote {len(feedback_records)} records")

        except Exception as e:
            self.logger.error(f"Error writing feedback records: {e}")
            raise

    def process(self, input_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        Main processing method for feedback generation.
        This is the interface method called by parent coordinator.

        Args:
            input_data: Optional dict with 'grade_records' or uses input file path

        Returns:
            Dictionary with feedback, output_file, generated_count, failed_count
        """
        self.logger.info("Starting feedback generation process")

        try:
            # Load grade records
            if input_data and 'grade_records' in input_data:
                grade_records = input_data['grade_records']
            else:
                input_file = self.config['input']['file_path']
                grade_records = self._load_grade_records(input_file)

            if not grade_records:
                self.logger.warning("No grade records to process")
                return {
                    'feedback': [],
                    'output_file': self.config['output']['file_path'],
                    'generated_count': 0,
                    'failed_count': 0
                }

            # Process each grade record
            feedback_records = []
            generated_count = 0
            failed_count = 0
            rate_limit_delay = self.config['rate_limiting']['delay_between_calls_seconds']

            for idx, record in enumerate(grade_records, 1):
                self.logger.info(f"Processing record {idx}/{len(grade_records)}: {record['email_id']}")

                feedback_record = self._generate_feedback_for_record(record)
                feedback_records.append(feedback_record)

                if feedback_record['status'] == 'Ready':
                    generated_count += 1
                else:
                    failed_count += 1

                # Apply rate limiting between calls (except for last record)
                if idx < len(grade_records) and rate_limit_delay > 0:
                    self.logger.debug(f"Waiting {rate_limit_delay}s before next request")
                    time.sleep(rate_limit_delay)

            # Write output file
            output_file = self.config['output']['file_path']
            self._write_feedback_records(feedback_records, output_file)

            # Return summary
            result = {
                'feedback': feedback_records,
                'output_file': output_file,
                'generated_count': generated_count,
                'failed_count': failed_count
            }

            self.logger.info(
                f"Process complete: {generated_count} generated, {failed_count} failed"
            )
            return result

        except Exception as e:
            self.logger.error(f"Error in feedback generation process: {e}")
            raise

    def health_check(self) -> Dict[str, Any]:
        """
        Perform health check on all services.

        Returns:
            Dictionary with health status of all components
        """
        health = {
            'feedback_manager': 'OK',
            'style_selector': 'Unknown',
            'gemini_generator': 'Unknown'
        }

        try:
            # Test style selector
            test_result = self.style_selector.select_style(75.0)
            if test_result and 'style_name' in test_result:
                health['style_selector'] = 'OK'
            else:
                health['style_selector'] = 'Failed'
        except Exception as e:
            health['style_selector'] = f'Error: {e}'

        try:
            # Test gemini generator (just config check, not API call)
            if hasattr(self.gemini_generator, 'model'):
                health['gemini_generator'] = 'OK (API not tested)'
            else:
                health['gemini_generator'] = 'Failed'
        except Exception as e:
            health['gemini_generator'] = f'Error: {e}'

        return health


def main():
    """Main entry point for standalone execution."""
    parser = argparse.ArgumentParser(
        description='Feedback Manager Service - Generate AI feedback for student grades'
    )
    parser.add_argument('--input', type=str,
                       help='Path to input Excel file with grades')
    parser.add_argument('--config', type=str, default='config.yaml',
                       help='Path to config file')
    parser.add_argument('--verbose', action='store_true',
                       help='Enable verbose logging')
    parser.add_argument('--health', action='store_true',
                       help='Run health check on all services')

    args = parser.parse_args()

    try:
        # Override log level if verbose
        if args.verbose:
            logging.getLogger().setLevel(logging.DEBUG)

        # Initialize manager
        manager = FeedbackManager(config_path=args.config)

        # Health check mode
        if args.health:
            print("\n" + "="*60)
            print("FEEDBACK MANAGER HEALTH CHECK")
            print("="*60)
            health = manager.health_check()
            for service, status in health.items():
                print(f"{service:.<30} {status}")
            print("="*60 + "\n")
            return 0

        # Process feedback generation
        input_data = None
        if args.input:
            # Override config input path
            manager.config['input']['file_path'] = args.input

        result = manager.process(input_data)

        # Display results
        print("\n" + "="*60)
        print("FEEDBACK GENERATION COMPLETE")
        print("="*60)
        print(f"Total records processed: {len(result['feedback'])}")
        print(f"Successfully generated:  {result['generated_count']}")
        print(f"Failed:                  {result['failed_count']}")
        print(f"Output file:             {result['output_file']}")
        print("="*60 + "\n")

        return 0

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
