"""
Grade Manager - Main Entry Point
Standalone execution for the grade manager service.
"""

import argparse
import sys
import os

# Add the parent directory to the path to allow imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from service import GradeManagerService


def main():
    """Main entry point for standalone execution."""
    parser = argparse.ArgumentParser(
        description='Grade Manager - Orchestrates repository grading'
    )
    parser.add_argument(
        '--input',
        help='Path to input file (file_1_2.xlsx with email records)',
        default=None
    )
    parser.add_argument(
        '--output',
        help='Path to output file (file_2_3.xlsx for grades)',
        default=None
    )
    parser.add_argument(
        '--config',
        default='config.yaml',
        help='Path to configuration file (default: config.yaml)'
    )
    parser.add_argument(
        '--workers',
        type=int,
        help='Number of parallel workers (overrides config)',
        default=None
    )
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Enable verbose output'
    )

    args = parser.parse_args()

    print("=" * 60)
    print("Grade Manager v1.0.0")
    print("=" * 60)
    print(f"Config file: {args.config}")

    if args.input:
        print(f"Input file: {args.input}")
    else:
        print("Input file: (none - health check mode)")

    if args.output:
        print(f"Output file: {args.output}")
    else:
        print("Output file: (from config)")

    if args.workers:
        print(f"Workers: {args.workers}")

    print("=" * 60)
    print()

    # Initialize service
    service = GradeManagerService(config_path=args.config)

    if args.input:
        # Process input file
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.input)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                print("Error: Input file is empty")
                return 1

            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            email_records = [
                dict(zip(headers, row)) for row in rows[1:] if any(row)
            ]

            print(f"Loaded {len(email_records)} email records")

            # Process records
            result = service.process({'email_records': email_records})

            print()
            print("=" * 60)
            print("Results:")
            print(f"  Successfully graded: {result['graded_count']}")
            print(f"  Failed: {result['failed_count']}")
            if result['output_file']:
                print(f"  Output file: {result['output_file']}")
            print("=" * 60)

            # Print individual results if verbose
            if args.verbose and result.get('grades'):
                print("\nDetailed Results:")
                for grade in result['grades']:
                    status_icon = "✓" if grade['status'] == 'Ready' else "✗"
                    print(f"  {status_icon} {grade['email_id'][:16]}... - Grade: {grade['grade']}")
                    if grade.get('error'):
                        print(f"      Error: {grade['error']}")

            return 0 if result['failed_count'] == 0 else 1

        except ImportError:
            print("Error: openpyxl is required. Install with: pip install openpyxl")
            return 1
        except FileNotFoundError:
            print(f"Error: Input file not found: {args.input}")
            return 1
        except Exception as e:
            print(f"Error: {e}")
            return 1
    else:
        # Health check mode
        health = service.health_check()
        print("Health Check:")
        print(f"  Manager: {health['manager']} - {health['status']}")
        for child, status in health.get('children', {}).items():
            status_icon = "✓" if status == "healthy" else "✗"
            print(f"  {status_icon} {child}: {status}")
        print()
        print("To grade repositories, provide an input file:")
        print(f"  python -m grade_manager --input ../../../email_coordinator/email_reader/data/output/file_1_2.xlsx")
        print()
        return 0


if __name__ == '__main__':
    sys.exit(main())
