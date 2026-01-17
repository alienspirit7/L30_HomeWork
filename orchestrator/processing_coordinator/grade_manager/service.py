"""
Grade Manager Service
Level 2 - Task Manager

Orchestrates repository grading by coordinating:
- GitHub Cloner: Clones repositories from GitHub URLs
- Python Analyzer: Analyzes Python files and calculates grades
"""

import sys
import os
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import yaml

# Add parent paths for imports
sys.path.insert(0, str(Path(__file__).parent))

# Import child services
try:
    from github_cloner.service import GitHubClonerService
except ImportError:
    GitHubClonerService = None

try:
    sys.path.insert(0, str(Path(__file__).parent / "python_analyzer"))
    from src.service import PythonAnalyzerService
except ImportError:
    PythonAnalyzerService = None

# Setup logging
logger = logging.getLogger(__name__)


class GradeManagerService:
    """
    Grade Manager Service - Level 2 Task Manager
    
    Orchestrates repository grading by coordinating child services:
    - GitHub Cloner: Clones repositories from GitHub
    - Python Analyzer: Analyzes Python code and calculates grades
    
    Input: List of email records with repository URLs
    Output: List of grade records with calculated grades
    """
    
    def __init__(self, config_path: str = "config.yaml"):
        """
        Initialize the Grade Manager Service.
        
        Args:
            config_path: Path to configuration file
        """
        self.config_path = config_path
        self.config = self._load_config(config_path)
        self._setup_logging()
        self._initialize_child_services()
        
        self.logger.info(f"Grade Manager Service v{self.config.get('manager', {}).get('version', '1.0.0')} initialized")
    
    def _load_config(self, config_path: str) -> Dict[str, Any]:
        """Load configuration from YAML file."""
        try:
            with open(config_path, 'r') as f:
                return yaml.safe_load(f) or {}
        except FileNotFoundError:
            logger.warning(f"Config file not found: {config_path}, using defaults")
            return self._default_config()
    
    def _default_config(self) -> Dict[str, Any]:
        """Return default configuration."""
        return {
            'manager': {
                'name': 'grade_manager',
                'version': '1.0.0'
            },
            'children': {
                'github_cloner': './github_cloner',
                'python_analyzer': './python_analyzer'
            },
            'output': {
                'file_path': './data/output/file_2_3.xlsx'
            },
            'parallelism': {
                'max_workers': 5
            },
            'cleanup': {
                'delete_repos_after_grading': True
            },
            'logging': {
                'level': 'INFO',
                'file': './logs/grade_manager.log'
            }
        }
    
    def _setup_logging(self) -> None:
        """Configure logging based on config settings."""
        log_config = self.config.get('logging', {})
        log_level = getattr(logging, log_config.get('level', 'INFO').upper())
        log_file = log_config.get('file', './logs/grade_manager.log')
        
        # Create logs directory
        log_path = Path(log_file)
        log_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Configure module logger
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(log_level)
        
        if not self.logger.handlers:
            # File handler
            fh = logging.FileHandler(log_file)
            fh.setLevel(log_level)
            fh.setFormatter(logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            ))
            self.logger.addHandler(fh)
            
            # Console handler
            ch = logging.StreamHandler()
            ch.setLevel(log_level)
            ch.setFormatter(logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s'
            ))
            self.logger.addHandler(ch)
    
    def _initialize_child_services(self) -> None:
        """Initialize child services (GitHub Cloner and Python Analyzer)."""
        children_config = self.config.get('children', {})
        
        # Get the directory where this service.py is located (NOT based on config path)
        service_dir = Path(__file__).parent.absolute()
        
        # Initialize GitHub Cloner
        if GitHubClonerService is not None:
            cloner_rel_path = children_config.get('github_cloner', './github_cloner')
            cloner_path = (service_dir / cloner_rel_path).resolve()
            cloner_config = cloner_path / 'config.yaml'
            try:
                # Change working directory temporarily for proper config resolution
                original_cwd = os.getcwd()
                os.chdir(cloner_path)
                self.github_cloner = GitHubClonerService(
                    config_path=str(cloner_config) if cloner_config.exists() else "config.yaml"
                )
                os.chdir(original_cwd)
                self.logger.info("GitHub Cloner service initialized")
            except Exception as e:
                self.logger.warning(f"Failed to initialize GitHub Cloner: {e}")
                self.github_cloner = None
                if 'original_cwd' in locals():
                    os.chdir(original_cwd)
        else:
            self.logger.warning("GitHubClonerService not available")
            self.github_cloner = None
        
        # Initialize Python Analyzer
        if PythonAnalyzerService is not None:
            analyzer_rel_path = children_config.get('python_analyzer', './python_analyzer')
            analyzer_path = (service_dir / analyzer_rel_path).resolve()
            analyzer_config = analyzer_path / 'config.yaml'
            try:
                original_cwd = os.getcwd()
                os.chdir(analyzer_path)
                self.python_analyzer = PythonAnalyzerService(
                    config_path=str(analyzer_config) if analyzer_config.exists() else "config.yaml"
                )
                os.chdir(original_cwd)
                self.logger.info("Python Analyzer service initialized")
            except Exception as e:
                self.logger.warning(f"Failed to initialize Python Analyzer: {e}")
                self.python_analyzer = None
                if 'original_cwd' in locals():
                    os.chdir(original_cwd)
        else:
            self.logger.warning("PythonAnalyzerService not available")
            self.python_analyzer = None
    
    def grade_single_repository(self, email_record: Dict[str, Any]) -> Dict[str, Any]:
        """
        Grade a single repository from an email record.
        
        Args:
            email_record: Dictionary containing email_id and repo_url
            
        Returns:
            Grade record with email_id, grade, and status
        """
        email_id = email_record.get('email_id', 'unknown')
        repo_url = email_record.get('repo_url', '')
        
        self.logger.info(f"Processing repository for email {email_id[:8]}...")
        
        result = {
            'email_id': email_id,
            'grade': 0.0,
            'status': 'Failed',
            'error': None
        }
        
        # Validate input
        if not repo_url:
            result['error'] = 'Missing repo_url'
            return result
        
        if self.github_cloner is None:
            result['error'] = 'GitHub Cloner service not available'
            return result
        
        if self.python_analyzer is None:
            result['error'] = 'Python Analyzer service not available'
            return result
        
        clone_path = None
        try:
            # Step 1: Clone repository
            clone_result = self.github_cloner.clone_repository(repo_url)
            
            if clone_result.get('status') != 'Success':
                result['error'] = f"Clone failed: {clone_result.get('error', 'Unknown error')}"
                return result
            
            clone_path = clone_result.get('clone_path')
            self.logger.info(f"Cloned to {clone_path}")
            
            # Step 2: Analyze Python files
            analysis_result = self.python_analyzer.analyze(clone_path)
            
            if analysis_result.get('status') == 'Failed':
                result['error'] = f"Analysis failed: {analysis_result.get('error', 'Unknown error')}"
                return result
            
            # Extract grade
            result['grade'] = round(analysis_result.get('grade', 0.0), 2)
            result['status'] = 'Ready'
            result['error'] = None
            
            self.logger.info(f"Grade for {email_id[:8]}: {result['grade']}")
            
        except Exception as e:
            self.logger.error(f"Error grading repository: {e}")
            result['error'] = str(e)
        
        finally:
            # Cleanup cloned repository
            if clone_path and self.config.get('cleanup', {}).get('delete_repos_after_grading', True):
                try:
                    if self.github_cloner:
                        self.github_cloner.cleanup_repository(clone_path)
                except Exception as e:
                    self.logger.warning(f"Cleanup failed: {e}")
        
        return result
    
    def process(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process all email records to generate grades.
        
        Args:
            input_data: Dictionary containing 'email_records' list
            
        Returns:
            Dictionary with grades, output_file, graded_count, failed_count
        """
        email_records = input_data.get('email_records', [])
        
        if not email_records:
            self.logger.warning("No email records to process")
            return {
                'grades': [],
                'output_file': None,
                'graded_count': 0,
                'failed_count': 0
            }
        
        # Filter to only "Ready" records
        ready_records = [
            r for r in email_records 
            if r.get('status') == 'Ready'
        ]
        
        self.logger.info(f"Processing {len(ready_records)} repositories (of {len(email_records)} total)")
        
        max_workers = self.config.get('parallelism', {}).get('max_workers', 5)
        grades = []
        
        # Process repositories (can be parallelized)
        if max_workers > 1 and len(ready_records) > 1:
            # Parallel processing
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {
                    executor.submit(self.grade_single_repository, record): record
                    for record in ready_records
                }
                
                for future in as_completed(futures):
                    try:
                        result = future.result()
                        grades.append(result)
                    except Exception as e:
                        record = futures[future]
                        grades.append({
                            'email_id': record.get('email_id', 'unknown'),
                            'grade': 0.0,
                            'status': 'Failed',
                            'error': str(e)
                        })
        else:
            # Sequential processing
            for record in ready_records:
                result = self.grade_single_repository(record)
                grades.append(result)
        
        # Calculate summary
        graded_count = sum(1 for g in grades if g.get('status') == 'Ready')
        failed_count = len(grades) - graded_count
        
        # Write to Excel if openpyxl is available
        output_file = None
        try:
            output_path = self.config.get('output', {}).get('file_path', './data/output/file_2_3.xlsx')
            self._write_grades_to_excel(grades, output_path)
            output_file = output_path
        except Exception as e:
            self.logger.warning(f"Failed to write Excel output: {e}")
        
        self.logger.info(f"Grading complete: {graded_count} successful, {failed_count} failed")
        
        return {
            'grades': grades,
            'output_file': output_file,
            'graded_count': graded_count,
            'failed_count': failed_count
        }
    
    def _write_grades_to_excel(self, grades: List[Dict], output_path: str) -> None:
        """Write grade records to Excel file."""
        try:
            from openpyxl import Workbook
        except ImportError:
            self.logger.warning("openpyxl not installed, skipping Excel output")
            return
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        
        # Header
        headers = ['email_id', 'grade', 'status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        for row_idx, grade in enumerate(grades, 2):
            ws.cell(row=row_idx, column=1, value=grade.get('email_id', ''))
            ws.cell(row=row_idx, column=2, value=grade.get('grade', 0))
            ws.cell(row=row_idx, column=3, value=grade.get('status', 'Failed'))
        
        wb.save(output_path)
        self.logger.info(f"Wrote {len(grades)} grades to {output_path}")
    
    def health_check(self) -> Dict[str, Any]:
        """Check health of manager and child services."""
        return {
            'manager': 'grade_manager',
            'status': 'healthy',
            'children': {
                'github_cloner': 'healthy' if self.github_cloner else 'unavailable',
                'python_analyzer': 'healthy' if self.python_analyzer else 'unavailable'
            }
        }


def main():
    """Main entry point for standalone execution."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Grade Manager Service')
    parser.add_argument('--input', help='Path to input Excel file (file_1_2.xlsx)')
    parser.add_argument('--output', help='Path to output Excel file (file_2_3.xlsx)')
    parser.add_argument('--config', default='config.yaml', help='Path to config file')
    parser.add_argument('--workers', type=int, help='Number of parallel workers')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Grade Manager Service v1.0.0")
    print("=" * 60)
    
    # Initialize service
    service = GradeManagerService(config_path=args.config)
    
    # Load input if provided
    if args.input:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(args.input)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            headers = rows[0]
            email_records = [
                dict(zip(headers, row)) for row in rows[1:] if any(row)
            ]
            
            print(f"Loaded {len(email_records)} records from {args.input}")
            
            # Process
            result = service.process({'email_records': email_records})
            
            print(f"\nResults:")
            print(f"  Graded: {result['graded_count']}")
            print(f"  Failed: {result['failed_count']}")
            if result['output_file']:
                print(f"  Output: {result['output_file']}")
            
        except Exception as e:
            print(f"Error: {e}")
            return 1
    else:
        # Health check only
        health = service.health_check()
        print(f"\nHealth Check:")
        print(f"  Manager: {health['manager']} - {health['status']}")
        for child, status in health.get('children', {}).items():
            print(f"  {child}: {status}")
    
    print("=" * 60)
    return 0


if __name__ == '__main__':
    sys.exit(main())
